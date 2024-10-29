import pandas as pd
import openpyxl
import os
import yaml
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import re

# Load config
with open('config.yaml', 'r') as file:
    config = yaml.safe_load(file)

def filter_sheets(input_file_path, output_file_path, po_numbers_to_exclude):
    with pd.ExcelFile(input_file_path) as xls:
        df_stack = pd.read_excel(xls, "Stack")
        df_pc_overview = pd.read_excel(xls, "pc_overview")

    df_stack_filtered = df_stack[~df_stack["PO #"].isin(po_numbers_to_exclude)]
    df_stack_removed = df_stack[df_stack["PO #"].isin(po_numbers_to_exclude)]

    df_pc_overview_filtered = df_pc_overview[~df_pc_overview["PO #"].isin(po_numbers_to_exclude)]
    df_pc_overview_removed = df_pc_overview[df_pc_overview["PO #"].isin(po_numbers_to_exclude)]

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        df_stack_filtered.to_excel(writer, sheet_name="Stack", index=False)
        df_pc_overview_filtered.to_excel(writer, sheet_name="pc_overview", index=False)
        df_stack_removed.to_excel(writer, sheet_name="Stack_Removed", index=False)
        df_pc_overview_removed.to_excel(writer, sheet_name="pc_overview_Removed", index=False)

def process_data():
    ALL_STACK_PATH = os.path.join("AP_Files", "All_Stack_Filtered.xlsx")
    OUTPUT_PATH = os.path.join("AP_Files", "updated_all_stack.xlsx")
    
    pc_overview = pd.read_excel(ALL_STACK_PATH, sheet_name="pc_overview")
    stack_data = pd.read_excel(ALL_STACK_PATH, sheet_name="Stack")
    
    new_rows = pc_overview[~pc_overview['PO #'].isin(stack_data['PO #'])]
    new_stack_rows = new_rows[['Project Number', 'PO #', 'PO Description', 'Vendor/Subcontractor', 'Amount']].copy()
    
    stack_data_updated = pd.concat([stack_data, new_stack_rows], ignore_index=True)
    
    with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
        stack_data_updated.to_excel(writer, sheet_name='pc_overview', index=False)

def update_stack_sheet():
    ALL_STACK_PATH = os.path.join("AP_Files", "All_Stack.xlsx")
    OUTPUT_PATH = os.path.join("AP_Files", "updated_all_stack.xlsx")
    
    summary_data = pd.read_excel(OUTPUT_PATH, sheet_name='pc_overview')
    stack_data = pd.read_excel(ALL_STACK_PATH, sheet_name='Stack')
    
    new_rows = summary_data[~summary_data['PO #'].isin(stack_data['PO #'])]
    
    if len(new_rows) > 0:
        with pd.ExcelWriter(ALL_STACK_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            updated_stack = pd.concat([stack_data, new_rows], ignore_index=True)
            updated_stack.to_excel(writer, sheet_name='Stack', index=False)

def update_ar_sheet():
    source_file = os.path.join("AR_Files", "AR_Analysis.xlsx")
    dest_file = os.path.join("AR_Files", "AR_updated.xlsx")
    
    updated_po_data = pd.read_excel(source_file, sheet_name='Updated PO Data')
    ar_data = pd.read_excel(dest_file, sheet_name='last_updated')
    
    updated_po_data['PO #'] = updated_po_data['PO #'].astype(str)
    ar_data['PO #'] = ar_data['PO #'].astype(str)
    
    new_rows = updated_po_data[~updated_po_data['PO #'].isin(ar_data['PO #'])]
    
    if len(new_rows) > 0:
        with pd.ExcelWriter(dest_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            updated_ar = pd.concat([ar_data, new_rows], ignore_index=True)
            updated_ar.to_excel(writer, sheet_name='last_updated', index=False)

def load_and_process_data():
    pc_overview = pd.read_excel(config['EXISTING_FILE_PATH'], sheet_name='pc_overview')
    updated_po_data = pd.read_excel(config['UPDATED_PO_DATA_PATH'], sheet_name="Updated PO Data")
    return pc_overview, updated_po_data

def generate_reports(pc_overview, updated_po_data):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    
    # Process and create Combined PM Types sheet
    combined_pm_types = process_pm_types(pc_overview, updated_po_data)
    
    # Write to Excel file
    with pd.ExcelWriter(config['NEW_FILE_PATH'], engine='openpyxl') as writer:
        # Write the main sheets
        pc_overview.to_excel(writer, sheet_name='pc_overview AP', index=False)
        updated_po_data.to_excel(writer, sheet_name='pc_overview AR', index=False)
        combined_pm_types.to_excel(writer, sheet_name='Combined PM Types', index=False)
        
        # Create detailed analysis
        detailed_pm_types = create_detailed_analysis(combined_pm_types)
        detailed_pm_types.to_excel(writer, sheet_name='Detailed Combined PM Types', index=False)
        
        # Create base-build breakdown
        base_build = create_base_build_breakdown(combined_pm_types)
        base_build.to_excel(writer, sheet_name='Base-Build_breakdown', index=False)

def process_pm_types(pc_overview, updated_po_data):
    # Combine and process the data
    combined_data = pd.concat([pc_overview, updated_po_data], ignore_index=True)
    
    # Add necessary calculations and columns
    combined_data['PM Type'] = combined_data['PO Description'].apply(categorize_pm_type)
    combined_data['Amount'] = pd.to_numeric(combined_data['Amount'], errors='coerce')
    
    # Group by PM Type and calculate totals
    pm_type_summary = combined_data.groupby('PM Type').agg({
        'Amount': 'sum',
        'PO #': 'count'
    }).reset_index()
    
    # Calculate percentages
    total_amount = pm_type_summary['Amount'].sum()
    pm_type_summary['Percentage'] = (pm_type_summary['Amount'] / total_amount) * 100
    
    return pm_type_summary

def categorize_pm_type(description):
    # Define PM Type categories based on PO Description
    description = str(description).lower()
    
    # Add your categorization logic here
    if 'mechanical' in description:
        return 'Mechanical'
    elif 'electrical' in description:
        return 'Electrical'
    elif 'plumbing' in description:
        return 'Plumbing'
    # Add more categories as needed
    else:
        return 'Other'

def create_detailed_analysis(combined_pm_types):
    # Create more detailed analysis of the PM Types
    detailed_analysis = combined_pm_types.copy()
    
    # Add additional analysis columns
    detailed_analysis['Running Total'] = detailed_analysis['Amount'].cumsum()
    detailed_analysis['Cumulative Percentage'] = detailed_analysis['Running Total'] / detailed_analysis['Amount'].sum() * 100
    
    return detailed_analysis

def create_base_build_breakdown(combined_pm_types):
    # Create base vs. build analysis
    base_build = combined_pm_types.copy()
    
    # Add base/build categorization
    base_build['Category'] = base_build['PM Type'].apply(categorize_base_build)
    
    # Group by category
    base_build_summary = base_build.groupby('Category').agg({
        'Amount': 'sum',
        'PO #': 'count'
    }).reset_index()
    
    return base_build_summary

def categorize_base_build(pm_type):
    # Define which PM Types are base vs. build
    base_types = ['Mechanical', 'Electrical', 'Plumbing']  # Add your base types
    return 'Base' if pm_type in base_types else 'Build'

def cleanup_workbook():
    wb = openpyxl.load_workbook(config['NEW_FILE_PATH'])
    sheets_to_keep = ['pc_overview AP', 'pc_overview AR', 'Combined PM Types', 'Detailed Combined PM Types', 'Base-Build_breakdown']
    for sheet_name in wb.sheetnames:
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]
    wb.save(config['NEW_FILE_PATH'])